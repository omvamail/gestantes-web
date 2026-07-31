"""
converter.py
Lógica de conversión de Excel → formato SIGIRES (MSPS).
Módulo reutilizable para la app web Flask.
"""
import re
import openpyxl
import datetime

SHEET_NAMES = [
    '1 - Control',
    '2 - ID gestantes',
    '3 - Atenciones',
    '4 - Seguimientos',
    '5 - Urgencias',
]


def clean_string(val):
    """Convierte a mayúsculas, elimina tildes y ñ."""
    if val is None:
        return ""
    val = str(val).upper().strip()
    for old, new in {'Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','Ü':'U','Ñ':'N'}.items():
        val = val.replace(old, new)
    return val


def clean_address(val):
    """
    Limpia el campo de dirección para SIGIRES.
    - Convierte a mayúsculas y elimina tildes/ñ
    - Elimina espacios INMEDIATAMENTE después del ;
      (el espacio después del ; no es válido según el formato)
    - Preserva espacios DENTRO de cada componente
      (ej: VDA;LAS PUERTAS  →  VDA;LAS PUERTAS  ✅)
    - Ejemplo incorrecto que corrige: VDA; LAS PUERTAS → VDA;LAS PUERTAS
    """
    if val is None:
        return ""
    val = clean_string(val)          # mayúsculas + sin tildes/ñ
    val = re.sub(r';\s+', ';', val)  # elimina espacios tras ;
    val = re.sub(r'\s+;', ';', val)  # elimina espacios antes de ;
    return val


def format_cell(cell):
    """Formatea una celda para la salida SIGIRES."""
    if cell is None:
        return ""
    if isinstance(cell, (datetime.datetime, datetime.date)):
        return cell.strftime("%Y-%m-%d")
    if isinstance(cell, datetime.time):
        return cell.strftime("%H:%M:%S")
    if isinstance(cell, float):
        return str(int(cell)) if cell.is_integer() else str(cell)
    return clean_string(cell)


def serialize_cell(cell):
    """Serializa una celda para JSON (preview)."""
    if cell is None:
        return None
    if isinstance(cell, (datetime.datetime, datetime.date)):
        return cell.strftime("%Y-%m-%d")
    if isinstance(cell, datetime.time):
        return cell.strftime("%H:%M:%S")
    if isinstance(cell, float):
        return int(cell) if cell.is_integer() else cell
    if isinstance(cell, (int, str, bool)):
        return cell
    return str(cell)


def get_num_cols(headers):
    """Cuenta columnas con encabezado no-None."""
    count = 0
    for h in headers:
        if h is not None:
            count += 1
        else:
            break
    return count


def read_excel_preview(file_path):
    """
    Lee el Excel y devuelve datos crudos para preview (JSON-serializable).
    Retorna: dict { nombre_hoja: [[...filas...]] }
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    preview = {}
    for name in SHEET_NAMES:
        if name in wb.sheetnames:
            sheet = wb[name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([serialize_cell(c) for c in row])
                if len(rows) >= 200:
                    break
            preview[name] = rows
        else:
            preview[name] = []
    return preview


def convert_to_sigires(file_path):
    """
    Convierte el Excel al formato de texto SIGIRES.
    Retorna: dict { filename, content, stats }
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    detail_lines = []
    ips_code = None
    end_date_str = None
    control_vals = []

    for idx, name in enumerate(SHEET_NAMES):
        if name not in wb.sheetnames:
            continue

        sheet = wb[name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        num_cols = get_num_cols(rows[0])

        for r in rows[1:]:
            if not any(cell is not None for cell in r):
                continue
            if r[0] != idx + 1:
                continue

            if idx > 0:
                if not any(
                    cell is not None and str(cell).strip() != ""
                    for cell in r[1:num_cols]
                ):
                    continue

            # Hoja 2: columna 17 = dirección → limpieza especial de separadores ;
            ADDRESS_COL = 17  # índice de "Dirección de residencia" en Tipo 2
            if idx == 1:
                formatted_row = []
                for ci, c in enumerate(r[:num_cols]):
                    if ci == ADDRESS_COL:
                        formatted_row.append(clean_address(c))
                    else:
                        formatted_row.append(format_cell(c))
            else:
                formatted_row = [format_cell(c) for c in r[:num_cols]]

            if idx == 0:
                control_vals = formatted_row
                if len(r) > 5 and r[5] is not None:
                    v = r[5]
                    if isinstance(v, (datetime.datetime, datetime.date)):
                        end_date_str = v.strftime("%d%m%Y")
                    else:
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                            try:
                                end_date_str = datetime.datetime.strptime(
                                    str(v).strip(), fmt
                                ).strftime("%d%m%Y")
                                break
                            except ValueError:
                                continue
            else:
                if idx == 1 and not ips_code and len(formatted_row) > 5:
                    ips_code = formatted_row[5]
                detail_lines.append("|".join(formatted_row))

    if control_vals:
        if len(control_vals) > 6:
            control_vals[6] = str(len(detail_lines))
        control_line = "|".join(control_vals)
    else:
        control_line = ""

    output_lines = ([control_line] if control_line else []) + detail_lines

    ips_code = ips_code or "DESCONOCIDO"
    end_date_str = end_date_str or datetime.datetime.now().strftime("%d%m%Y")
    filename = f"GESTANTE_MSPS_{ips_code}_{end_date_str}.txt"

    return {
        "filename": filename,
        "content": "\n".join(output_lines) + "\n",
        "stats": {
            "filename": filename,
            "ips_code": ips_code,
            "end_date": end_date_str,
            "total_detail": len(detail_lines),
        },
    }
